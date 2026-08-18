import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from bot.admin_auth import is_admin
from database.database import Database
from services.admin import broadcast_copy, refresh_all_vip_balances
from services.excel_export import VIP_EXPORT_HEADERS, build_vip_excel


class FakeOurbitAPI:
    def __init__(self, balances):
        self.balances = balances

    async def get_balance(self, uid):
        return self.balances.get(uid)


class FakeBot:
    def __init__(self, failing_user=None):
        self.failing_user = failing_user
        self.sent = []

    async def copy_message(
        self,
        chat_id,
        from_chat_id,
        message_id,
    ):
        if chat_id == self.failing_user:
            raise RuntimeError("delivery failed")
        self.sent.append((chat_id, from_chat_id, message_id))
        return True


class AdminAuthorizationTest(unittest.TestCase):
    def test_admin_ids_are_explicit(self):
        allowed = frozenset({100, 200})
        self.assertTrue(is_admin(100, allowed))
        self.assertFalse(is_admin(300, allowed))
        self.assertFalse(is_admin(None, allowed))


class AdminDatabaseTest(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = str(Path(self.temp_dir.name) / "admin-test.db")
        self.db = Database(self.db_path)
        await self.db.init()

    async def asyncTearDown(self):
        self.temp_dir.cleanup()

    async def test_stats_search_tracking_and_recipients(self):
        await self.db.track_bot_user(1, "visitor", "Visitor")
        status = await self.db.register_user(
            telegram_id=2,
            username="vip",
            first_name="VIP",
            ourbit_uid="12345678",
            balance=25,
            invite_link="https://example.invalid/invite",
        )
        self.assertEqual(status, "created")

        stats = await self.db.get_admin_stats()
        self.assertEqual(stats["total_bot_users"], 2)
        self.assertEqual(stats["active_vips"], 1)
        self.assertEqual(stats["inactive_vips"], 0)

        by_uid = await self.db.find_user("12345678")
        by_telegram = await self.db.find_user("2")
        self.assertEqual(by_uid["telegram_id"], 2)
        self.assertEqual(by_telegram["ourbit_uid"], "12345678")

        recipients = await self.db.get_broadcast_user_ids()
        self.assertEqual(recipients, [1, 2])

        tracked_user = await self.db.get_bot_user(1)
        self.assertEqual(tracked_user["username"], "visitor")

        all_vips = await self.db.get_all_vip_users()
        self.assertEqual(len(all_vips), 1)
        self.assertEqual(all_vips[0]["telegram_id"], 2)

    async def test_live_refresh_updates_without_enforcement(self):
        await self.db.register_user(
            telegram_id=10,
            username=None,
            first_name="Low",
            ourbit_uid="11111111",
            balance=20,
            invite_link="link-1",
        )
        await self.db.register_user(
            telegram_id=20,
            username=None,
            first_name="Safe",
            ourbit_uid="22222222",
            balance=20,
            invite_link="link-2",
        )
        api = FakeOurbitAPI(
            {
                "11111111": 5.0,
                "22222222": 15.0,
            }
        )

        result = await refresh_all_vip_balances(
            minimum_balance=10,
            database=self.db,
            api=api,
            concurrency=2,
        )

        self.assertEqual(result.checked, 2)
        self.assertEqual(result.failed, 0)
        self.assertEqual(result.below_minimum, 1)
        low_user = await self.db.get_user(10)
        self.assertEqual(low_user["balance"], 5.0)
        self.assertEqual(low_user["vip_status"], "active")
        self.assertIsNone(low_user["last_check"])


class AdminBroadcastTest(unittest.IsolatedAsyncioTestCase):
    async def test_broadcast_isolates_delivery_failures(self):
        bot = FakeBot(failing_user=2)
        result = await broadcast_copy(
            bot=bot,
            user_ids=[1, 2, 3, 3],
            from_chat_id=99,
            message_id=7,
        )

        self.assertEqual(result.total, 3)
        self.assertEqual(result.sent, 2)
        self.assertEqual(result.failed, 1)
        self.assertEqual(
            bot.sent,
            [(1, 99, 7), (3, 99, 7)],
        )


class AdminExcelExportTest(unittest.TestCase):
    def test_excel_contains_username_and_vip_fields(self):
        users = [
            {
                "username": "vip_user",
                "first_name": "VIP",
                "telegram_id": 123,
                "ourbit_uid": "12345678",
                "vip_status": "active",
                "balance": 20.5,
                "joined_at": "2026-08-18 10:00:00",
                "last_check": "2026-08-18 11:00:00",
                "last_warning": None,
                "warning_count": 1,
            }
        ]

        content = build_vip_excel(users)
        workbook = load_workbook(BytesIO(content), read_only=True)
        sheet = workbook["VIP Users"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(list(rows[0]), VIP_EXPORT_HEADERS)
        self.assertEqual(rows[1][0], "@vip_user")
        self.assertEqual(rows[1][3], "12345678")
        self.assertEqual(rows[1][4], "active")
        self.assertEqual(rows[1][5], 20.5)


if __name__ == "__main__":
    unittest.main()
